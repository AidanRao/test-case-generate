import io
import json
import os
import tempfile
import unittest
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.models.testcase import PRIORITY_LEVELS, SCENARIO_TYPES
from testcase_generator import TestCaseGenerator, _build_system_prompt


class TestCaseGeneratorScenarioTest(unittest.TestCase):
    def _generator_with_response(self, payload):
        client = MagicMock()
        client.chat.completions.create.return_value = SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(content=json.dumps(payload, ensure_ascii=False))
                )
            ]
        )
        return TestCaseGenerator(client, "test-model")

    def test_prompt_contains_all_scenarios_and_applicability_rule(self):
        prompt = _build_system_prompt("功能测试")

        for scenario_type in SCENARIO_TYPES:
            self.assertIn(scenario_type, prompt)
        self.assertIn("尽量覆盖所有适用的场景类别", prompt)
        self.assertIn("不适用的类别可以不生成", prompt)
        self.assertNotIn("priority", prompt)

    def test_accepts_valid_scenario(self):
        generator = self._generator_with_response(
            [{"title": "正常登录", "scenario_type": "正常流程用例", "test_steps": []}]
        )

        result = generator.generate_test_cases("允许用户登录", "REQ-1", "登录")

        self.assertEqual(result[0]["scenario_type"], "正常流程用例")
        self.assertEqual(result[0]["test_case_type"], "功能测试")

    def test_rejects_missing_or_invalid_scenario(self):
        invalid_payloads = [
            [{"title": "未分类", "test_steps": []}],
            [{"title": "未知分类", "scenario_type": "其他用例", "test_steps": []}],
        ]

        for payload in invalid_payloads:
            with self.subTest(payload=payload):
                generator = self._generator_with_response(payload)
                self.assertIsNone(
                    generator.generate_test_cases("允许用户登录", "REQ-1", "登录")
                )


class _StubGenerator:
    def __init__(self, cases):
        self.cases = cases

    def generate_test_cases(self, *args, **kwargs):
        return self.cases


class TestCaseScenarioApiTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmpdir.cleanup)
        self.environment = patch.dict(
            os.environ,
            {"DATA_DIR": self.tmpdir.name, "UNIPORTAL_SYNC_ENABLED": "false"},
            clear=False,
        )
        self.environment.start()
        self.addCleanup(self.environment.stop)

        from app import create_app

        self.app = create_app()
        self.app.testing = True
        self.client = self.app.test_client()
        response = self.client.post(
            "/v1/projects",
            json={
                "code": "PRJ-SCENARIO",
                "title": "场景测试项目",
                "requirements": [
                    {
                        "module": "登录",
                        "requirements": [
                            {
                                "title": "用户登录",
                                "type": "功能需求",
                                "code": "REQ-LOGIN",
                                "content": "用户输入账号密码后登录",
                            }
                        ],
                    }
                ],
            },
        )
        self.project_id = response.get_json()["data"]["id"]
        self.requirement = self.app.config["STORAGE"].list_requirements(
            self.project_id
        )[0]
        self.testcase_id = "TC-SCENARIO-001"
        self.app.config["STORAGE"].add_testcases(
            self.project_id,
            self.requirement["id"],
            [
                {
                    "id": self.testcase_id,
                    "requirement_code": self.requirement["code"],
                    "requirement_id": self.requirement["id"],
                    "title": "验证正常登录",
                    "code": "TC-PRJ-SCENARIO-001",
                    "type": "功能测试",
                    "scenario_type": "正常流程用例",
                    "test_steps": [
                        {"step_desc": "输入正确账号密码", "expectation": "登录成功"}
                    ],
                    "test_target_desc": "验证正常登录流程",
                    "verify_method": "TESTING",
                }
            ],
        )
        self.testcases_path = os.path.join(self.tmpdir.name, "testcases.json")
        with open(self.testcases_path, "r", encoding="utf-8") as testcase_file:
            stored_testcases = json.load(testcase_file)
        stored_testcases[0].pop("priority", None)
        with open(self.testcases_path, "w", encoding="utf-8") as testcase_file:
            json.dump(stored_testcases, testcase_file, ensure_ascii=False, indent=2)

    def _update_payload(self, scenario_type):
        return {
            "title": "验证登录",
            "code": "TC-PRJ-SCENARIO-001",
            "type": "功能测试",
            "scenario_type": scenario_type,
            "priority": "P1",
            "test_steps": [
                {"step_desc": "输入账号密码", "expectation": "返回登录结果"}
            ],
            "test_target_desc": "验证登录流程",
            "verify_method": "TESTING",
        }

    def test_updates_and_returns_every_supported_scenario(self):
        for scenario_type in SCENARIO_TYPES:
            with self.subTest(scenario_type=scenario_type):
                response = self.client.put(
                    f"/v1/projects/{self.project_id}/testcases/{self.testcase_id}",
                    json=self._update_payload(scenario_type),
                )
                self.assertEqual(response.status_code, 200)
                listed = self.client.get(
                    f"/v1/projects/{self.project_id}/requirements/"
                    f"{self.requirement['id']}/testcases"
                ).get_json()["data"]["list"]
                self.assertEqual(listed[0]["scenario_type"], scenario_type)

        detail = self.client.get(f"/v1/projects/{self.project_id}").get_json()["data"]
        self.assertEqual(
            detail["requirements"][0]["testcases"][0]["scenario_type"],
            SCENARIO_TYPES[-1],
        )

    def test_rejects_missing_and_invalid_scenario(self):
        missing = self._update_payload("正常流程用例")
        missing.pop("scenario_type")

        for payload in (missing, self._update_payload("其他用例")):
            with self.subTest(payload=payload):
                response = self.client.put(
                    f"/v1/projects/{self.project_id}/testcases/{self.testcase_id}",
                    json=payload,
                )
                self.assertEqual(response.status_code, 400)
                self.assertEqual(response.get_json()["code"], 40001)

    def test_defaults_priority_to_p1_and_updates_all_supported_levels(self):
        with open(self.testcases_path, "r", encoding="utf-8") as testcase_file:
            stored_testcases = json.load(testcase_file)
        self.assertNotIn("priority", stored_testcases[0])

        listed = self.client.get(
            f"/v1/projects/{self.project_id}/requirements/"
            f"{self.requirement['id']}/testcases"
        ).get_json()["data"]["list"]
        self.assertEqual(listed[0]["priority"], "P1")

        for priority in PRIORITY_LEVELS:
            with self.subTest(priority=priority):
                payload = self._update_payload("正常流程用例")
                payload["priority"] = priority
                response = self.client.put(
                    f"/v1/projects/{self.project_id}/testcases/{self.testcase_id}",
                    json=payload,
                )
                self.assertEqual(response.status_code, 200)
                listed = self.client.get(
                    f"/v1/projects/{self.project_id}/requirements/"
                    f"{self.requirement['id']}/testcases"
                ).get_json()["data"]["list"]
                self.assertEqual(listed[0]["priority"], priority)
                with open(self.testcases_path, "r", encoding="utf-8") as testcase_file:
                    stored_testcases = json.load(testcase_file)
                self.assertEqual(stored_testcases[0]["priority"], priority)

    def test_rejects_invalid_priority(self):
        payload = self._update_payload("正常流程用例")
        payload["priority"] = "P4"

        response = self.client.put(
            f"/v1/projects/{self.project_id}/testcases/{self.testcase_id}",
            json=payload,
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.get_json()["code"], 40001)

    def test_excel_export_contains_scenario_and_preserves_step_layout(self):
        self.client.put(
            f"/v1/projects/{self.project_id}/testcases/{self.testcase_id}",
            json=self._update_payload("组合场景用例"),
        )

        response = self.client.get(f"/v1/projects/{self.project_id}/testcases/export")
        self.addCleanup(response.close)

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data))
        self.addCleanup(workbook.close)
        sheet = workbook.active
        labels = {
            sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value
            for row in range(1, 8)
        }
        self.assertEqual(labels["用例场景"], "组合场景用例")
        self.assertEqual(sheet.cell(row=7, column=1).value, "测试步骤")
        self.assertEqual(sheet.cell(row=8, column=1).value, 1)
        self.assertEqual(sheet.cell(row=8, column=2).value, "输入账号密码")

    def test_integration_markdown_contains_scenario(self):
        generator = _StubGenerator(
            [
                {
                    "title": "异常登录",
                    "test_case_type": "功能测试",
                    "scenario_type": "异常场景用例",
                    "test_steps": [
                        {"step_desc": "输入错误密码", "expectation": "提示登录失败"}
                    ],
                    "test_target_desc": "验证异常登录",
                    "verify_method": "TESTING",
                }
            ]
        )
        payload = {
            "format": "md",
            "is_save": False,
            "requirements": [
                {
                    "module": "登录",
                    "requirements": [
                        {
                            "id": "REQ-MD",
                            "title": "用户登录",
                            "type": "功能需求",
                            "code": "REQ-MD",
                            "content": "用户输入账号密码后登录",
                        }
                    ],
                }
            ],
        }

        with patch(
            "app.services.testcase_service.TestCaseService._build_generator",
            return_value=generator,
        ):
            response = self.client.post("/v1/integration/testcases/generate", json=payload)

        self.assertEqual(response.status_code, 200)
        self.assertIn("- 用例场景：异常场景用例", response.get_json()["test_case"])


if __name__ == "__main__":
    unittest.main()
