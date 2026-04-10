from openpyxl import load_workbook
from copy import copy

class TestCaseExporter:
    def __init__(self, template_path):
        self.template_path = template_path
        self.wb = load_workbook(template_path)
        self.template_ws = self.wb.active
        # Pre-process template information
        self._analyze_template()

    def _analyze_template(self):
        # Dynamically analyze template structure
        self.meta_rows = 0
        self.step_start_row = 0
        self.meta_field_map = {} # Map label to row index (relative to 1)

        # Field Mapping Config
        # Label in Excel -> Key in JSON
        self.label_to_key = {
            '测试用例标识': 'code',
            '测试目标描述': 'test_target_desc',
            '被测需求标识': 'requirement_code',
            '验证方法': 'verify_method',
            '测试用例类型': 'test_case_type'
        }

        # 1. Find the Header Row (containing "测试步骤")
        # Scan first 20 rows
        for r in range(1, 21):
            cell_val = self.template_ws.cell(row=r, column=1).value
            # Check column 1 or any column for specific header keywords
            # Based on previous knowledge, "测试步骤" is in the header row, maybe col 1
            # But let's look for the row that splits Metadata and Steps.
            # Usually the Header Row has '测试步骤', '测试输入及测试步骤', '预期结果' etc.
            
            # Let's check all cells in this row
            row_values = [self.template_ws.cell(row=r, column=c).value for c in range(1, 6)]
            if any(isinstance(v, str) and '测试步骤' in v for v in row_values):
                self.header_row_index = r
                self.step_start_row = r + 1
                self.meta_rows = r - 1
                break
        
        if self.step_start_row == 0:
            # Fallback if not found (should not happen with correct template)
            self.header_row_index = 6
            self.step_start_row = 7
            self.meta_rows = 5

        # 2. Map Metadata Rows
        for r in range(1, self.meta_rows + 1):
            # Check the label in the first column (or merged cells starting in first col)
            label_cell = self.template_ws.cell(row=r, column=1)
            label = str(label_cell.value).strip() if label_cell.value else ""
            
            # Match label to key
            for known_label, key in self.label_to_key.items():
                if known_label in label:
                    self.meta_field_map[r] = key
                    break

        # Capture row heights
        self.row_heights = {}
        for r in range(1, self.header_row_index + 2): # +1 for step row (start row)
            if r in self.template_ws.row_dimensions:
                 self.row_heights[r] = self.template_ws.row_dimensions[r].height

        # Capture merged cells in the metadata block
        self.meta_merged_cells = []
        for merged_range in self.template_ws.merged_cells.ranges:
            if merged_range.min_row <= self.meta_rows:
                self.meta_merged_cells.append(merged_range)

    def _copy_cell_style(self, src_cell, target_cell):
        if src_cell.has_style:
            target_cell.font = copy(src_cell.font)
            target_cell.border = copy(src_cell.border)
            target_cell.fill = copy(src_cell.fill)
            target_cell.number_format = src_cell.number_format
            target_cell.protection = copy(src_cell.protection)
            target_cell.alignment = copy(src_cell.alignment)

    def _reset_sheet(self, target_ws):
        for merged_range in list(target_ws.merged_cells.ranges):
            try:
                target_ws.unmerge_cells(str(merged_range))
            except Exception:
                pass
        max_r = target_ws.max_row
        max_c = target_ws.max_column
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = target_ws.cell(row=r, column=c)
                cell.value = None
                cell.style = 'Normal'

    def _render_cases(self, target_ws, test_cases):
        current_row = 1
        
        for tc in test_cases:
            start_row = current_row
            
            # 1. Write Metadata Block (Rows 1 to meta_rows)
            for r_offset in range(self.meta_rows):
                src_row = r_offset + 1
                tgt_row = current_row + r_offset
                
                # Copy Row Height
                if src_row in self.row_heights:
                    target_ws.row_dimensions[tgt_row].height = self.row_heights[src_row]

                for col in range(1, self.template_ws.max_column + 1):
                    src_cell = self.template_ws.cell(row=src_row, column=col)
                    tgt_cell = target_ws.cell(row=tgt_row, column=col)
                    
                    self._copy_cell_style(src_cell, tgt_cell)
                    tgt_cell.value = src_cell.value

                    # Fill Data Dynamically based on Label Map
                    # Usually value is in Column 2 (B)
                    if col == 2:
                        # Check if this row is mapped to a key
                        if src_row in self.meta_field_map:
                            key = self.meta_field_map[src_row]
                            
                            if key == 'test_proc_id':
                                tc_id = tc.get('test_case_id', '')
                                tgt_cell.value = tc_id.replace('TC-', 'TP-') if 'TC-' in tc_id else tc_id
                            else:
                                tgt_cell.value = tc.get(key, '')

            # Apply Merged Cells for Metadata Block
            for merged_range in self.meta_merged_cells:
                # Only apply merges that fall within the metadata rows
                if merged_range.min_row <= self.meta_rows:
                    min_row = merged_range.min_row + (current_row - 1)
                    max_row = merged_range.max_row + (current_row - 1)
                    min_col = merged_range.min_col
                    max_col = merged_range.max_col
                    
                    target_ws.merge_cells(start_row=min_row, start_column=min_col, 
                                        end_row=max_row, end_column=max_col)

            current_row += self.meta_rows
            
            # 2. Write Step Header (Header Row)
            src_header_row = self.header_row_index
            tgt_header_row = current_row
            if src_header_row in self.row_heights:
                target_ws.row_dimensions[tgt_header_row].height = self.row_heights[src_header_row]
            
            for col in range(1, self.template_ws.max_column + 1):
                src_cell = self.template_ws.cell(row=src_header_row, column=col)
                tgt_cell = target_ws.cell(row=tgt_header_row, column=col)
                self._copy_cell_style(src_cell, tgt_cell)
                tgt_cell.value = src_cell.value
            
            # Apply Merged Cells for Header Row
            for merged_range in self.template_ws.merged_cells.ranges:
                if merged_range.min_row == src_header_row and merged_range.max_row == src_header_row:
                     target_ws.merge_cells(start_row=tgt_header_row, start_column=merged_range.min_col, 
                                         end_row=tgt_header_row, end_column=merged_range.max_col)

            current_row += 1

            # 3. Write Steps
            steps = tc.get('test_steps', [])
            step_start_row = current_row
            src_step_row = self.step_start_row # The row in template that defines step style
            
            # If no steps, we might still want one empty row to show structure? 
            # Assuming at least 1 step based on logic.
            if not steps:
                steps = [{}] # Dummy step
                
            for i, step in enumerate(steps, 1):
                tgt_row = current_row
                # Copy Row Height from template step row
                if src_step_row in self.row_heights:
                    target_ws.row_dimensions[tgt_row].height = self.row_heights[src_step_row]
                
                for col in range(1, self.template_ws.max_column + 1):
                    src_cell = self.template_ws.cell(row=src_step_row, column=col) # Always copy from Row 7
                    tgt_cell = target_ws.cell(row=tgt_row, column=col)
                    self._copy_cell_style(src_cell, tgt_cell)
                    
                    # Fill Data
                    if col == 1: # Step Num
                        tgt_cell.value = i
                    elif col == 2: # Step Desc
                        tgt_cell.value = step.get('step_desc', '')
                    elif col == 3: # Expectation
                        tgt_cell.value = step.get('expectation', '')
                    elif col == 4: # Criteria
                        # If it's the first step, set the criteria text
                        if i == 1:
                            tgt_cell.value = "实测结果与预期结果一致或再预期结果范围内，则测试通过；反之，测试失败"
                        else:
                            tgt_cell.value = None
                          
                current_row += 1
                
            # 4. Handle Right-Side Merged Column (Criteria)
            # Merge Col 4 from step_start_row to (current_row - 1)
            # This covers exactly the step rows.
            if steps:
                # Merge logic:
                # Start: step_start_row
                # End: current_row - 1 (the last step row)
                # Col: 4
                target_ws.merge_cells(start_row=step_start_row, start_column=4, 
                                    end_row=current_row - 1, end_column=4)
 
            # 5. Empty Row / Separator
            # We don't copy style for the empty row, just leave it standard
            current_row += 1

    def _safe_sheet_title(self, title, used):
        invalid = ['\\', '/', '*', '[', ']', ':', '?']
        safe = ''.join(ch for ch in str(title) if ch not in invalid).strip()
        if not safe:
            safe = "Sheet"
        safe = safe[:31]
        candidate = safe
        index = 1
        while candidate in used:
            suffix = f"-{index}"
            candidate = f"{safe[:31 - len(suffix)]}{suffix}"
            index += 1
        used.add(candidate)
        return candidate

    def export(self, test_cases, output_path):
        target_ws = self.wb.copy_worksheet(self.template_ws)
        target_ws.title = "Exported Test Cases"
        self._reset_sheet(target_ws)
        self._render_cases(target_ws, test_cases)
        # Remove the original template sheet
        self.wb.remove(self.template_ws)
        self.wb.save(output_path)

    def export_by_requirement(self, groups, output_path):
        used = set()
        if not groups:
            groups = [{"sheet_name": "Empty", "test_cases": []}]
        for group in groups:
            target_ws = self.wb.copy_worksheet(self.template_ws)
            title = self._safe_sheet_title(group.get("sheet_name", "Sheet"), used)
            target_ws.title = title
            self._reset_sheet(target_ws)
            self._render_cases(target_ws, group.get("test_cases", []))
        self.wb.remove(self.template_ws)
        self.wb.save(output_path)


if __name__ == '__main__':
    tce = TestCaseExporter("template/test_case_export_template.xlsx")
